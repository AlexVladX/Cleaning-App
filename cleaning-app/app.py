import os
import json
import io
import threading
import time
import urllib.request
import urllib.error
import queue
from datetime import datetime
from flask import Flask, render_template, jsonify, send_file, request
from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload
import pdfplumber
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from collections import defaultdict

app = Flask(__name__)

GROQ_API_KEY = os.environ.get("GROQ_API_KEY", "").strip()
GOOGLE_CREDENTIALS_JSON = os.environ.get("GOOGLE_CREDENTIALS_JSON", "")
FOLDER_ID = os.environ.get("DRIVE_FOLDER_ID", "")
POLL_INTERVAL = int(os.environ.get("POLL_INTERVAL", "60"))

state = {
    "files": {},
    "interventii": [],
    "last_scan": None,
    "monitoring": False,
    "log": []
}
state_lock = threading.Lock()
file_queue = queue.Queue()


def add_log(msg):
    ts = datetime.now().strftime("%H:%M:%S")
    with state_lock:
        state["log"].append(f"[{ts}] {msg}")
        if len(state["log"]) > 200:
            state["log"] = state["log"][-200:]


def get_drive_service():
    if not GOOGLE_CREDENTIALS_JSON:
        raise ValueError("GOOGLE_CREDENTIALS_JSON not set")
    creds_dict = json.loads(GOOGLE_CREDENTIALS_JSON)
    creds = service_account.Credentials.from_service_account_info(
        creds_dict,
        scopes=["https://www.googleapis.com/auth/drive.readonly"]
    )
    return build("drive", "v3", credentials=creds)


def list_pdf_files(folder_id):
    service = get_drive_service()
    all_files = []

    def search_folder(fid):
        results = service.files().list(
            q=f"'{fid}' in parents and trashed=false",
            fields="files(id, name, mimeType, modifiedTime, size)",
            pageSize=100
        ).execute()
        for f in results.get("files", []):
            if f["mimeType"] == "application/pdf":
                all_files.append(f)
            elif f["mimeType"] == "application/vnd.google-apps.folder":
                add_log(f"Subfolder: {f['name']}")
                search_folder(f["id"])

    search_folder(folder_id)
    return all_files


def download_pdf(file_id):
    service = get_drive_service()
    req = service.files().get_media(fileId=file_id)
    buf = io.BytesIO()
    downloader = MediaIoBaseDownload(buf, req)
    done = False
    while not done:
        _, done = downloader.next_chunk()
    buf.seek(0)
    return buf


def extract_text_from_pdf(pdf_bytes):
    text = ""
    with pdfplumber.open(pdf_bytes) as pdf:
        for page in pdf.pages:
            tables = page.extract_tables()
            if tables:
                for table in tables:
                    for row in table:
                        if row:
                            clean = [str(c).strip() if c else "" for c in row]
                            text += " | ".join(clean) + "\n"
            else:
                t = page.extract_text()
                if t:
                    text += t + "\n"
    return text.strip()


def analyze_with_gemini(text, filename):
    json_example = '[{"proprietate":"Antoniotti","client":"Conciergerie La Clef Dor","data":"03/04/2026","tip_serviciu":"Standard","tip_facturare":"forfait","qty":1,"pret_unitar":117.0,"suma":117.0,"moneda":"EUR"}]'

    prompt = (
        "Esti contabil pentru o firma de curatenie. Analizeaza acest tabel extras dintr-o factura.\n\n"
        "STRUCTURA FACTURII:\n"
        "- In ANTET gasesti 'Facture pour' urmat de numele CLIENTULUI (firma de conciergerie sau persoana care plateste)\n"
        "- In TABELUL 'Detail des prestations' gasesti coloanele: Nom | Type | Date de nettoyage | Qty/heures | Prix unitaire | Prix total\n"
        "- Coloana 'Nom' = numele CASEI/PROPRIETATII (ex: ANTONIOTTI, FLAMENT, Villa Marina) - NU clientul!\n\n"
        "REGULI:\n"
        "1. proprietate = valoarea din coloana 'Nom' din tabel (casa, vila, apartamentul)\n"
        "2. client = valoarea de dupa 'Facture pour' din antet (firma sau persoana care plateste)\n"
        "3. O casa cu qty=3 si 3 date diferite = 3 interventii separate, cate una per data\n"
        "4. Imparte suma totala egal: ex 351 EUR / 3 interventii = 117 EUR fiecare\n"
        "5. Normalizeaza numele casei: Title Case. Ex: ANTONIOTTI => Antoniotti\n"
        "6. Data format: DD/MM/YYYY (converteste din MM/DD/YYYY)\n"
        "7. Daca pret unitar < 40 EUR => tip_facturare=heure\n"
        "   Daca pret unitar >= 40 EUR => tip_facturare=forfait\n\n"
        "Raspunde DOAR cu array JSON valid, fara explicatii, fara markdown.\n"
        "Format exemplu: " + json_example + "\n\n"
        "Tabelul facturii:\n" + text[:5000]
    )
    url = "https://api.groq.com/openai/v1/chat/completions"
    body = json.dumps({
        "model": "llama-3.1-8b-instant",
        "messages": [{"role": "user", "content": prompt}],
        "max_tokens": 2000
    }).encode()
    req = urllib.request.Request(
        url, data=body,
        headers={
            "Content-Type": "application/json",
            "Authorization": f"Bearer {GROQ_API_KEY.strip()}",
            "User-Agent": "Cleaning-App/1.0"
        }
    )
    try:
        with urllib.request.urlopen(req, timeout=60) as resp:
            data = json.loads(resp.read())
    except urllib.error.HTTPError as e:
        if e.code == 429:
            add_log("Rate limit. Astept 20 secunde si reincerc...")
            time.sleep(20)
            return analyze_with_gemini(text, filename)
        raise
    raw = data["choices"][0]["message"]["content"].strip()
    raw = raw.replace("```json", "").replace("```", "").strip()
    s = raw.find("[")
    e = raw.rfind("]") + 1
    if s >= 0 and e > s:
        interventii = json.loads(raw[s:e])
    else:
        s = raw.find("{")
        e = raw.rfind("}") + 1
        interventii = json.loads(raw[s:e]).get("interventii", [])
    for iv in interventii:
        if iv.get("proprietate"):
            iv["proprietate"] = " ".join(
                w.capitalize() for w in iv["proprietate"].strip().split()
            )
    return {"interventii": interventii}



def process_file(file_info):
    file_id = file_info["id"]
    name = file_info["name"]
    add_log(f"Descarc: {name}")
    with state_lock:
        state["files"][file_id] = {
            "id": file_id, "name": name,
            "status": "processing", "interventii": [], "error": None
        }
    try:
        pdf_bytes = download_pdf(file_id)
        add_log(f"Extrag text: {name}")
        text = extract_text_from_pdf(pdf_bytes)
        if not text:
            raise ValueError("PDF scanat sau gol")
        add_log(f"Analizez cu AI: {name}")
        result = analyze_with_gemini(text, name)
        interventii = result.get("interventii", [])
        for iv in interventii:
            iv["_source"] = name
            iv["_file_id"] = file_id
        with state_lock:
            state["files"][file_id]["status"] = "done"
            state["files"][file_id]["interventii"] = interventii
            state["interventii"] = [
                iv for iv in state["interventii"] if iv.get("_file_id") != file_id
            ] + interventii
        add_log(f"✓ {name}: {len(interventii)} intervenții extrase")
    except Exception as e:
        with state_lock:
            state["files"][file_id]["status"] = "error"
            state["files"][file_id]["error"] = str(e)
        add_log(f"✗ Eroare la {name}: {e}")


def process_queue_background(new_files):
    """Procesează lista de fișiere unul câte unul în background."""
    add_log(f"Background: incep procesarea a {len(new_files)} fisiere...")
    for i, f in enumerate(new_files):
        add_log(f"[{i+1}/{len(new_files)}] Procesez: {f['name']}")
        process_file(f)
        if i < len(new_files) - 1:
            add_log("Pauza 15s...")
            time.sleep(15)
    add_log("Toate fisierele procesate!")


def scan_folder():
    folder_id = FOLDER_ID
    if not folder_id:
        add_log("DRIVE_FOLDER_ID nu e setat")
        return
    add_log("Scanez folderul Drive...")
    try:
        files = list_pdf_files(folder_id)
        add_log(f"{len(files)} PDF-uri gasite in total")
        with state_lock:
            known_ids = set(state["files"].keys())
        new_files = [f for f in files if f["id"] not in known_ids]
        if new_files:
            add_log(f"{len(new_files)} fisiere noi - pornesc procesarea in background")
            for f in new_files:
                with state_lock:
                    state["files"][f["id"]] = {
                        "id": f["id"], "name": f["name"],
                        "status": "new", "interventii": [], "error": None
                    }
            t = threading.Thread(target=process_queue_background, args=(new_files,))
            t.daemon = True
            t.start()
            add_log(f"Thread background pornit (id={t.ident})")
        else:
            add_log("Nicio factura noua.")
        with state_lock:
            state["last_scan"] = datetime.now().isoformat()
    except Exception as e:
        import traceback
        add_log(f"Eroare scanare: {e} | {traceback.format_exc()[-300:]}")


def monitor_loop():
    while True:
        with state_lock:
            monitoring = state["monitoring"]
        if monitoring:
            scan_folder()
        time.sleep(POLL_INTERVAL)


def build_excel():
    with state_lock:
        interventii = list(state["interventii"])
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Clasament proprietati"
    prop_map = defaultdict(lambda: {"count": 0, "total": 0.0, "client": ""})
    for iv in interventii:
        k = iv.get("proprietate") or "Necunoscut"
        prop_map[k]["count"] += 1
        prop_map[k]["total"] += float(iv.get("suma") or 0)
        if iv.get("client"):
            prop_map[k]["client"] = iv["client"]
    props = sorted(prop_map.items(), key=lambda x: x[1]["count"], reverse=True)
    max_count = props[0][1]["count"] if props else 1
    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB")
    )
    headers1 = ["Rang", "Proprietate", "Client", "Interventii", "Venit total (EUR)", "Performanta"]
    ws1.append(headers1)
    for col in range(1, 7):
        cell = ws1.cell(1, col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin
    for i, (name, data) in enumerate(props, 1):
        perf = ("Performant" if data["count"] >= max_count * 0.7
                else "Mediu" if data["count"] >= max_count * 0.4 else "Rar")
        ws1.append([i, name, data["client"], data["count"], round(data["total"], 2), perf])
        for col in range(1, 7):
            cell = ws1.cell(i + 1, col)
            cell.border = thin
            cell.alignment = Alignment(horizontal="center" if col in [1, 4, 6] else "left")
            if perf == "Performant":
                cell.fill = PatternFill("solid", fgColor="D1FAE5")
            elif perf == "Rar":
                cell.fill = PatternFill("solid", fgColor="FEE2E2")
    for col, w in zip("ABCDEF", [6, 30, 20, 14, 18, 14]):
        ws1.column_dimensions[col].width = w
    ws2 = wb.create_sheet("Detalii interventii")
    headers2 = ["Proprietate", "Client", "Data", "Tip serviciu", "Suma (EUR)", "Moneda", "Sursa factura"]
    ws2.append(headers2)
    for col in range(1, 8):
        cell = ws2.cell(1, col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin
    for iv in interventii:
        ws2.append([
            iv.get("proprietate", ""), iv.get("client", ""), iv.get("data", ""),
            iv.get("tip_serviciu", ""), round(float(iv.get("suma") or 0), 2),
            iv.get("moneda", "EUR"), iv.get("_source", "")
        ])
        for col in range(1, 8):
            ws2.cell(ws2.max_row, col).border = thin
    for col, w in zip("ABCDEFG", [28, 20, 12, 28, 12, 10, 28]):
        ws2.column_dimensions[col].width = w
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


@app.route("/")
def index():
    return render_template("index.html")


@app.route("/api/state")
def api_state():
    with state_lock:
        prop_map = defaultdict(lambda: {"forfait": 0, "ore": 0.0, "total": 0.0, "client": ""})
        for iv in state["interventii"]:
            k = iv.get("proprietate") or "Necunoscut"
            if iv.get("type_facturation") == "heure":
                prop_map[k]["ore"] += float(iv.get("qty") or 0)
            else:
                prop_map[k]["forfait"] += 1
            prop_map[k]["total"] += float(iv.get("suma") or 0)
            if iv.get("client"):
                prop_map[k]["client"] = iv["client"]
        max_c = max((v["forfait"] for v in prop_map.values()), default=1)
        if max_c == 0:
            max_c = 1
        s = {
            "files": list(state["files"].values()),
            "last_scan": state["last_scan"],
            "monitoring": state["monitoring"],
            "log": state["log"][-50:],
            "metrics": {
                "files": len(state["files"]),
                "analyzed": sum(1 for f in state["files"].values() if f["status"] == "done"),
                "interventii": len(state["interventii"]),
                "revenue": round(sum(float(iv.get("suma") or 0) for iv in state["interventii"]), 2)
            },
            "props": sorted([
                {
                    "name": k,
                    "forfait": v["forfait"],
                    "ore": round(v["ore"], 1),
                    "total": round(v["total"], 2),
                    "client": v["client"],
                    "perf": ("Performant" if v["forfait"] >= max_c * 0.7
                             else "Mediu" if v["forfait"] >= max_c * 0.4 else "Rar"),
                    "pct": round(v["forfait"] / max_c * 100)
                }
                for k, v in prop_map.items()
            ], key=lambda x: x["forfait"], reverse=True)
        }
    return jsonify(s)


@app.route("/api/scan", methods=["POST"])
def api_scan():
    t = threading.Thread(target=scan_folder)
    t.daemon = True
    t.start()
    return jsonify({"ok": True})


@app.route("/api/monitor", methods=["POST"])
def api_monitor():
    data = request.json or {}
    with state_lock:
        state["monitoring"] = data.get("active", False)
    add_log("Monitorizare " + ("activata" if state["monitoring"] else "oprita"))
    return jsonify({"monitoring": state["monitoring"]})


@app.route("/api/export")
def api_export():
    buf = build_excel()
    date_str = datetime.now().strftime("%Y-%m-%d")
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"analiza-curatenie-{date_str}.xlsx"
    )


@app.route("/api/test")
def api_test():
    return jsonify({
        "groq_key": bool(GROQ_API_KEY),
        "folder_id": bool(FOLDER_ID),
        "queue_size": file_queue.qsize(),
        "files_count": len(state["files"]),
        "worker_alive": True
    })


@app.route("/api/reset", methods=["POST"])
def api_reset():
    with state_lock:
        state["files"] = {}
        state["interventii"] = []
        state["last_scan"] = None
        state["monitoring"] = False
        state["log"] = []
    add_log("Reset complet.")
    return jsonify({"ok": True})


if __name__ == "__main__":
    monitor_thread = threading.Thread(target=monitor_loop)
    monitor_thread.daemon = True
    monitor_thread.start()
    add_log("Aplicatie pornita. Procesare seriala cu pauza 15s intre facturi.")
    if FOLDER_ID:
        scan_folder()
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=False)



@app.route("/api/test-groq")
def test_groq():
    req = urllib.request.Request(
        "https://api.groq.com/openai/v1/models",
        headers={
            "Authorization": f"Bearer {GROQ_API_KEY.strip()}"
        }
    )

    try:
        with urllib.request.urlopen(req) as r:
            return r.read().decode()
    except urllib.error.HTTPError as e:
        return e.read().decode(), e.code

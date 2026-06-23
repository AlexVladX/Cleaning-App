import os
import json
import io
import threading
import time
import urllib.request
from datetime import datetime
from flask import Flask, render_template, jsonify, send_file, request
from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload
import pdfplumber
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from collections import defaultdict

app = Flask(__name__)

# --- Config ---
GEMINI_API_KEY = os.environ.get("GEMINI_API_KEY", "")
GOOGLE_CREDENTIALS_JSON = os.environ.get("GOOGLE_CREDENTIALS_JSON", "")
FOLDER_ID = os.environ.get("DRIVE_FOLDER_ID", "")
POLL_INTERVAL = int(os.environ.get("POLL_INTERVAL", "60"))

# --- State ---
state = {
    "files": {},
    "interventii": [],
    "last_scan": None,
    "monitoring": False,
    "log": []
}
state_lock = threading.Lock()


def add_log(msg):
    ts = datetime.now().strftime("%H:%M:%S")
    with state_lock:
        state["log"].append(f"[{ts}] {msg}")
        if len(state["log"]) > 200:
            state["log"] = state["log"][-200:]


def get_drive_service():
    if not GOOGLE_CREDENTIALS_JSON:
        raise ValueError("GOOGLE_CREDENTIALS_JSON not set")
    creds_dict = json.loads(GOOGLE_CREDENTIALS_JSON)
    creds = service_account.Credentials.from_service_account_info(
        creds_dict,
        scopes=["https://www.googleapis.com/auth/drive.readonly"]
    )
    return build("drive", "v3", credentials=creds)


def list_pdf_files(folder_id):
    service = get_drive_service()
    all_files = []

    def search_folder(fid):
        results = service.files().list(
            q=f"'{fid}' in parents and trashed=false",
            fields="files(id, name, mimeType, modifiedTime, size)",
            pageSize=100
        ).execute()
        for f in results.get("files", []):
            if f["mimeType"] == "application/pdf":
                all_files.append(f)
            elif f["mimeType"] == "application/vnd.google-apps.folder":
                add_log(f"Intru in subfolder: {f['name']}")
                search_folder(f["id"])

    search_folder(folder_id)
    return all_files


def download_pdf(file_id):
    service = get_drive_service()
    req = service.files().get_media(fileId=file_id)
    buf = io.BytesIO()
    downloader = MediaIoBaseDownload(buf, req)
    done = False
    while not done:
        _, done = downloader.next_chunk()
    buf.seek(0)
    return buf


def extract_text_from_pdf(pdf_bytes):
    text = ""
    with pdfplumber.open(pdf_bytes) as pdf:
        for page in pdf.pages:
            t = page.extract_text()
            if t:
                text += t + "\n"
    return text.strip()


def analyze_with_gemini(text, filename):
    prompt = f"""Ești contabil pentru o firmă de curățenie din Corsica. Analizează această factură și extrage toate intervențiile de curățenie.

FIȘIER: {filename}
CONȚINUT:
{text[:4000]}

Returnează DOAR un obiect JSON valid (fără markdown, fără backtick-uri):
{{
  "interventii": [
    {{
      "proprietate": "Numele vilei/casei/apartamentului",
      "adresa": "adresa dacă există",
      "data": "DD/MM/YYYY",
      "tip_serviciu": "tip de curățenie",
      "suma": 150.00,
      "moneda": "EUR",
      "client": "numele clientului dacă există"
    }}
  ]
}}

Dacă nu găsești date clare, deduce din context. Returnează lista goală dacă nu e o factură de curățenie."""

    url = f"https://generativelanguage.googleapis.com/v1beta/models/gemini-1.5-flash:generateContent?key={GEMINI_API_KEY}"
    body = json.dumps({"contents": [{"parts": [{"text": prompt}]}]}).encode()
    req = urllib.request.Request(url, data=body, headers={"Content-Type": "application/json"})
    with urllib.request.urlopen(req, timeout=30) as resp:
        data = json.loads(resp.read())
    raw = data["candidates"][0]["content"]["parts"][0]["text"].strip()
    raw = raw.replace("```json", "").replace("```", "").strip()
    return json.loads(raw)


def process_file(file_info):
    file_id = file_info["id"]
    name = file_info["name"]
    add_log(f"Descarc: {name}")

    with state_lock:
        state["files"][file_id] = {
            "id": file_id, "name": name,
            "status": "processing", "interventii": [], "error": None
        }

    try:
        pdf_bytes = download_pdf(file_id)
        add_log(f"Extrag text: {name}")
        text = extract_text_from_pdf(pdf_bytes)

        if not text:
            raise ValueError("PDF scanat sau gol — nu s-a putut extrage text")

        add_log(f"Analizez cu AI: {name}")
        result = analyze_with_gemini(text, name)
        interventii = result.get("interventii", [])
        for iv in interventii:
            iv["_source"] = name
            iv["_file_id"] = file_id

        with state_lock:
            state["files"][file_id]["status"] = "done"
            state["files"][file_id]["interventii"] = interventii
            state["interventii"] = [
                iv for iv in state["interventii"] if iv.get("_file_id") != file_id
            ] + interventii

        add_log(f"✓ {name}: {len(interventii)} intervenții extrase")

    except Exception as e:
        with state_lock:
            state["files"][file_id]["status"] = "error"
            state["files"][file_id]["error"] = str(e)
        add_log(f"✗ Eroare la {name}: {e}")


def scan_folder():
    folder_id = FOLDER_ID
    if not folder_id:
        add_log("DRIVE_FOLDER_ID nu e setat")
        return

    add_log("Scanez folderul Drive (inclusiv subdosare)...")
    try:
        files = list_pdf_files(folder_id)
        add_log(f"{len(files)} PDF-uri găsite în total")

        with state_lock:
            known_ids = set(state["files"].keys())

        new_files = [f for f in files if f["id"] not in known_ids]
        if new_files:
            add_log(f"{len(new_files)} fișiere noi de analizat")
            for f in new_files:
                t = threading.Thread(target=process_file, args=(f,))
                t.daemon = True
                t.start()
        else:
            add_log("Nicio factură nouă.")

        with state_lock:
            state["last_scan"] = datetime.now().isoformat()

    except Exception as e:
        add_log(f"Eroare scanare: {e}")


def monitor_loop():
    while True:
        with state_lock:
            monitoring = state["monitoring"]
        if monitoring:
            scan_folder()
        time.sleep(POLL_INTERVAL)


def build_excel():
    with state_lock:
        interventii = list(state["interventii"])

    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Clasament proprietăți"

    prop_map = defaultdict(lambda: {"count": 0, "total": 0.0, "client": ""})
    for iv in interventii:
        k = iv.get("proprietate") or "Necunoscut"
        prop_map[k]["count"] += 1
        prop_map[k]["total"] += float(iv.get("suma") or 0)
        if iv.get("client"):
            prop_map[k]["client"] = iv["client"]

    props = sorted(prop_map.items(), key=lambda x: x[1]["count"], reverse=True)
    max_count = props[0][1]["count"] if props else 1

    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB")
    )

    headers1 = ["Rang", "Proprietate", "Client", "Intervenții", "Venit total (€)", "Performanță"]
    ws1.append(headers1)
    for col, h in enumerate(headers1, 1):
        cell = ws1.cell(1, col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin

    for i, (name, data) in enumerate(props, 1):
        perf = "Performant" if data["count"] >= max_count * 0.7 else "Mediu" if data["count"] >= max_count * 0.4 else "Rar"
        row = [i, name, data["client"], data["count"], round(data["total"], 2), perf]
        ws1.append(row)
        for col in range(1, 7):
            cell = ws1.cell(i + 1, col)
            cell.border = thin
            cell.alignment = Alignment(horizontal="center" if col in [1, 4, 6] else "left")
            if perf == "Performant":
                cell.fill = PatternFill("solid", fgColor="D1FAE5")
            elif perf == "Rar":
                cell.fill = PatternFill("solid", fgColor="FEE2E2")

    ws1.column_dimensions["A"].width = 6
    ws1.column_dimensions["B"].width = 30
    ws1.column_dimensions["C"].width = 20
    ws1.column_dimensions["D"].width = 14
    ws1.column_dimensions["E"].width = 18
    ws1.column_dimensions["F"].width = 14

    ws2 = wb.create_sheet("Detalii intervenții")
    headers2 = ["Proprietate", "Client", "Data", "Tip serviciu", "Sumă (€)", "Monedă", "Sursă factură"]
    ws2.append(headers2)
    for col, h in enumerate(headers2, 1):
        cell = ws2.cell(1, col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin

    for iv in interventii:
        row = [
            iv.get("proprietate", ""),
            iv.get("client", ""),
            iv.get("data", ""),
            iv.get("tip_serviciu", ""),
            round(float(iv.get("suma") or 0), 2),
            iv.get("moneda", "EUR"),
            iv.get("_source", "")
        ]
        ws2.append(row)
        r = ws2.max_row
        for col in range(1, 8):
            ws2.cell(r, col).border = thin

    ws2.column_dimensions["A"].width = 28
    ws2.column_dimensions["B"].width = 20
    ws2.column_dimensions["C"].width = 12
    ws2.column_dimensions["D"].width = 28
    ws2.column_dimensions["E"].width = 12
    ws2.column_dimensions["F"].width = 10
    ws2.column_dimensions["G"].width = 28

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


@app.route("/")
def index():
    return render_template("index.html")


@app.route("/api/state")
def api_state():
    with state_lock:
        s = {
            "files": list(state["files"].values()),
            "total_interventii": len(state["interventii"]),
            "last_scan": state["last_scan"],
            "monitoring": state["monitoring"],
            "log": state["log"][-50:],
            "metrics": {
                "files": len(state["files"]),
                "analyzed": sum(1 for f in state["files"].values() if f["status"] == "done"),
                "interventii": len(state["interventii"]),
                "revenue": round(sum(float(iv.get("suma") or 0) for iv in state["interventii"]), 2)
            },
            "props": []
        }
        prop_map = defaultdict(lambda: {"count": 0, "total": 0.0})
        for iv in state["interventii"]:
            k = iv.get("proprietate") or "Necunoscut"
            prop_map[k]["count"] += 1
            prop_map[k]["total"] += float(iv.get("suma") or 0)
        max_c = max((v["count"] for v in prop_map.values()), default=1)
        s["props"] = sorted([
            {
                "name": k,
                "count": v["count"],
                "total": round(v["total"], 2),
                "perf": "Performant" if v["count"] >= max_c * 0.7 else "Mediu" if v["count"] >= max_c * 0.4 else "Rar",
                "pct": round(v["count"] / max_c * 100)
            }
            for k, v in prop_map.items()
        ], key=lambda x: x["count"], reverse=True)
    return jsonify(s)


@app.route("/api/scan", methods=["POST"])
def api_scan():
    t = threading.Thread(target=scan_folder)
    t.daemon = True
    t.start()
    return jsonify({"ok": True})


@app.route("/api/monitor", methods=["POST"])
def api_monitor():
    data = request.json or {}
    with state_lock:
        state["monitoring"] = data.get("active", False)
    add_log("Monitorizare " + ("activată" if state["monitoring"] else "oprită"))
    return jsonify({"monitoring": state["monitoring"]})


@app.route("/api/export")
def api_export():
    buf = build_excel()
    date_str = datetime.now().strftime("%Y-%m-%d")
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"analiza-curatenie-{date_str}.xlsx"
    )


@app.route("/api/reset", methods=["POST"])
def api_reset():
    with state_lock:
        state["files"] = {}
        state["interventii"] = []
        state["last_scan"] = None
        state["monitoring"] = False
        state["log"] = []
    add_log("Reset complet.")
    return jsonify({"ok": True})


if __name__ == "__main__":
    monitor_thread = threading.Thread(target=monitor_loop)
    monitor_thread.daemon = True
    monitor_thread.start()
    add_log("Aplicație pornită.")
    if FOLDER_ID:
        scan_folder()
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=False)
